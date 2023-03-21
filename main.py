import time
from datetime import date
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import os
# import webdriver
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains

def creat_excel():
    workbook = Workbook()
    path=os.getcwd()+'/Currencies/Currencies.xlsx'

    sheet = workbook.active
    sheet["A1"] = "Symbol"
    sheet['A1'].font = Font(bold=True)
    sheet["B1"] = "Name"
    sheet['B1'].font = Font(bold=True)
    sheet["C1"] = "Last Price"
    sheet['C1'].font = Font(bold=True)
    sheet["D1"] = "Change"
    sheet['D1'].font = Font(bold=True)
    sheet["E1"] = "% Change"
    sheet['E1'].font = Font(bold=True)
    sheet["D1"] = "Change"
    sheet['D1'].font = Font(bold=True)
    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=10)

    sheet.column_dimensions = dim_holder

    workbook.save(path)

def scarpyy():
    # create webdriver object
    chrome_options = webdriver.ChromeOptions()
    prefs = {'download.default_directory' : os.getcwd()+'/Currencies'}
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(os.getcwd()+'/chromedriver',chrome_options=chrome_options)
    
    
    # get link
    driver.get(' https://finance.yahoo.com/')
    
    # get element 
    element = driver.find_element("link text","Markets")
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    element = driver.find_element("link text","Currencies")
    actions.click(on_element = element).perform()
    time.sleep(8)
    
    #we get all html elements at once
    allelements=driver.page_source
    
    start=allelements.find('aria-label="Name"')
    allelements=allelements[start:start+60000]
    
    
    #%%Extract Currencies table
    Names=[]
    symbols=[]
    prices=[]
    changes=[]
    Changes=[]
    while True:
    
        x1=allelements.find('aria-label="Name"')
        if x1<0:
            break
        x2=allelements[x1+18:].find('</')
        Names.append(allelements[x1+18:x1+18+x2])
        x1=allelements.find('symbol="')
        x2=allelements[x1+8:].find('=')
        symbols.append(allelements[x1+8:x1+8+x2])
        x1=allelements.find('regularMarketPrice')
        x2=allelements[x1+18:].find('value="')+x1+18
        x1=allelements[x2+7:].find('">')
        prices.append(float(allelements[x2+7:x2+7+x1]))
        x1=allelements.find('regularMarketChange')
        x2=allelements[x1+19:].find('value="')+x1+19
        x1=allelements[x2+7:].find('">')
        changes.append(float(allelements[x2+7:x2+7+x1]))
        x1=allelements.find('regularMarketChangePercent')
        x2=allelements[x1+26:].find('value="')+x1+26
        x1=allelements[x2+7:].find('">')
        Changes.append(float(allelements[x2+7:x2+7+x1]))
        allelements=allelements[x2+1000:]
        
    #%% write table to xlsx

    creat_excel()
    wb=load_workbook( os.getcwd()+'/Currencies/Currencies.xlsx')
    today = date.today()
    year = today.year
    sn=wb.sheetnames
    d1 = today.strftime("%d-%m-%Y")
    ss_sheet1= wb[sn[0]]
    ss_sheet1.title =d1
    ws = wb.active
    
    for i in range(len(Names)):
        ws.cell(row=2+i,column=1,value=symbols[i]) 
        ws.cell(row=2+i,column=2,value=Names[i]) 
        ws.cell(row=2+i,column=3,value=prices[i]) 
        ws.cell(row=2+i,column=4,value=changes[i]) 
        ws.cell(row=2+i,column=5,value=Changes[i]) 
        #red color
        if Changes[i]<-0.1:
            ws.cell(2+i,5).fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
        elif  Changes[i]>-0.1 and Changes[i]<=0:
            ws.cell(2+i,5).fill=PatternFill(start_color="00FF6600", end_color="00FF6600", fill_type = "solid")
        elif   Changes[i]<0.1 and  Changes[i]>0:
            ws.cell(2+i,5).fill=PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type = "solid")
        elif   Changes[i]>0.1:
            ws.cell(2+i,5).fill=PatternFill(start_color="0000FF00", end_color="0000FF00", fill_type = "solid")
    wb.save(os.getcwd()+'/Currencies/Currencies.xlsx')
    wb.close() 
    
    #%% clicl change
    for i in range(1,6):
        element = driver.find_element("xpath",'//*[@id="list-res-table"]/div[1]/table/thead/tr/th[5]')
        actions.click(on_element = element).perform()
        actions.click(on_element = element).perform()
        
        
        element = driver.find_element("xpath",'//*[@id="list-res-table"]/div[1]/table/tbody/tr['+str(i)+']/td[1]/a')
        actions.click(on_element = element).perform()
        time.sleep(5)
        
        
            
        try:
            element = driver.find_element("xpath",'//*[@id="myLightboxContainer"]/section/button[2]')
            actions.click(on_element = element).perform()
            time.sleep(2)
        except:
            pass
        
        element = driver.find_element("link text","Historical Data")
        actions.click(on_element = element).perform()
        time.sleep(2)
        
        element = driver.find_element("xpath",'//*[@id="Col1-1-HistoricalDataTable-Proxy"]/section/div[1]/div[1]/div[1]/div/div/div')
        actions.click(on_element = element).perform()
        time.sleep(2)
        
        element = driver.find_element("xpath",'//*[@id="dropdown-menu"]/div/div[2]/input')
        element.send_keys(today.strftime("%d.%m.%Y"))
        
        element = driver.find_element("xpath",'//*[@id="dropdown-menu"]/div/div[1]/input')
        element.send_keys(today.strftime('01'+".%m."+str(year-1)))
        
        
        element = driver.find_element("xpath",'//*[@id="dropdown-menu"]/div/div[3]/button[1]')
        actions.click(on_element = element).perform()
        
        
        
        element = driver.find_element("xpath",'//*[@id="Col1-1-HistoricalDataTable-Proxy"]/section/div[1]/div[1]/button')
        actions.click(on_element = element).perform()
        time.sleep(2)
        element = driver.find_element("xpath",'//*[@id="Col1-1-HistoricalDataTable-Proxy"]/section/div[1]/div[2]/span[2]/a')
        actions.click(on_element = element).perform()
        time.sleep(5)
        
        driver.back()
        
        driver.back()
        
        driver.back()
        time.sleep(5)
    driver.quit()
if __name__ == "__main__":
    scarpyy()